import csv
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO, StringIO

from django.urls import reverse
from openpyxl import load_workbook

from apps.billing.tests.base import BillingAPITestCase
from apps.core.roles import Role
from apps.core.tenancy import tenant_context
from apps.properties.models import Bed, Room
from apps.residents.models import Resident

TODAY = date.today()
PERIOD = {
    'period_start': TODAY.isoformat(),
    'period_end': (TODAY + timedelta(days=30)).isoformat(),
    'due_date': (TODAY + timedelta(days=10)).isoformat(),
}


def _csv_rows(response):
    return list(csv.reader(StringIO(response.content.decode())))


def _xlsx_rows(response):
    workbook = load_workbook(BytesIO(response.content))
    return [[cell.value for cell in row] for row in workbook.active.iter_rows()]


class ExportTestCase(BillingAPITestCase):
    def setUp(self):
        super().setUp()
        self.tenant = self.create_tenant()
        self.owner = self.create_owner(self.tenant)
        self.manager = self.create_manager(self.tenant)
        self.receptionist = self.create_receptionist(self.tenant)
        self.property = self.create_property(self.tenant, name='Sunrise PG')
        self.floor = self.create_floor(self.property)
        self.room = self.create_room(
            self.floor, room_number='101', sharing_type=Room.SharingType.FOUR, category=Room.Category.AC,
            rack_rate_with_food=Decimal('7000.00'), rack_rate_without_food=Decimal('5500.00'),
        )
        self.bed = self.create_bed(self.room, bed_number='101-A')
        self.authenticate(self.owner)

    def _generate_and_issue_invoice(self, resident, **overrides):
        payload = {**PERIOD, **overrides, 'resident': str(resident.id)}
        generated = self.client.post(reverse('invoice-list'), payload).data
        self.client.post(reverse('invoice-issue', args=[generated['id']]))
        return generated['id']

    def _pay(self, invoice_id, amount, payment_date):
        return self.client.post(reverse('payment-list'), {
            'invoice': invoice_id, 'amount': str(amount),
            'payment_date': payment_date.isoformat(), 'payment_mode': 'upi',
        })


class ResidentExportTests(ExportTestCase):
    def setUp(self):
        super().setUp()
        self.resident = self.create_resident(
            self.property, first_name='Ravi', last_name='Kumar', phone='9000000001',
            status=Resident.Status.RESERVED,
        )
        self.check_in(self.resident, self.bed)

    def test_csv_export_contains_resident_row(self):
        response = self.client.get(reverse('export-residents'), {'filetype': 'csv'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')
        rows = _csv_rows(response)
        self.assertEqual(rows[0], ['First Name', 'Last Name', 'Phone', 'Email', 'Property', 'Status', 'Room / Bed'])
        self.assertIn(['Ravi', 'Kumar', '9000000001', '', 'Sunrise PG', 'Active', '101 (101-A)'], rows)

    def test_xlsx_export_contains_resident_row(self):
        response = self.client.get(reverse('export-residents'), {'filetype': 'xlsx'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        rows = _xlsx_rows(response)
        self.assertIn(['Ravi', 'Kumar', '9000000001', None, 'Sunrise PG', 'Active', '101 (101-A)'], rows)

    def test_pdf_export_returns_a_valid_pdf(self):
        response = self.client.get(reverse('export-residents'), {'filetype': 'pdf'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF'))

    def test_invalid_filetype_is_rejected(self):
        response = self.client.get(reverse('export-residents'), {'filetype': 'json'})
        self.assertEqual(response.status_code, 400)

    def test_default_format_is_csv(self):
        response = self.client.get(reverse('export-residents'))
        self.assertEqual(response['Content-Type'], 'text/csv')

    def test_receptionist_can_export_residents(self):
        self.authenticate(self.receptionist)
        response = self.client.get(reverse('export-residents'))
        self.assertEqual(response.status_code, 200)

    def test_manager_sees_only_assigned_properties(self):
        other_property = self.create_property(self.tenant, name='Other PG')
        other_floor = self.create_floor(other_property)
        other_room = self.create_room(other_floor, room_number='201')
        other_bed = self.create_bed(other_room, bed_number='201-A')
        other_resident = self.create_resident(
            other_property, first_name='Suresh', phone='9000000002', status=Resident.Status.RESERVED,
        )
        self.check_in(other_resident, other_bed)
        self.assign_staff(self.manager, self.property)  # not assigned to other_property

        self.authenticate(self.manager)
        rows = _csv_rows(self.client.get(reverse('export-residents')))

        names = [row[0] for row in rows[1:]]
        self.assertIn('Ravi', names)
        self.assertNotIn('Suresh', names)


class PaymentExportTests(ExportTestCase):
    def setUp(self):
        super().setUp()
        self.resident = self.create_resident(self.property, status=Resident.Status.RESERVED)
        self.check_in(self.resident, self.bed)
        self.invoice_id = self._generate_and_issue_invoice(self.resident)

    def test_csv_export_contains_payment_row(self):
        self._pay(self.invoice_id, Decimal('2000.00'), TODAY)

        rows = _csv_rows(self.client.get(reverse('export-payments')))
        amounts = [row[4] for row in rows[1:]]
        self.assertIn('2000.00', amounts)

    def test_receptionist_is_forbidden(self):
        self.authenticate(self.receptionist)
        response = self.client.get(reverse('export-payments'))
        self.assertEqual(response.status_code, 403)

    def test_manager_is_allowed(self):
        self.assign_staff(self.manager, self.property)
        self.authenticate(self.manager)
        response = self.client.get(reverse('export-payments'))
        self.assertEqual(response.status_code, 200)


class OutstandingDuesExportTests(ExportTestCase):
    def setUp(self):
        super().setUp()
        self.resident = self.create_resident(self.property, status=Resident.Status.RESERVED)
        self.check_in(self.resident, self.bed)

    def test_fully_paid_invoice_excluded(self):
        invoice_id = self._generate_and_issue_invoice(self.resident)
        self._pay(invoice_id, Decimal('7000.00'), TODAY)

        rows = _csv_rows(self.client.get(reverse('export-outstanding-dues')))
        self.assertEqual(len(rows), 1)  # header only

    def test_unpaid_past_due_invoice_marked_overdue(self):
        past_due = (TODAY - timedelta(days=5)).isoformat()
        self._generate_and_issue_invoice(self.resident, due_date=past_due)

        rows = _csv_rows(self.client.get(reverse('export-outstanding-dues')))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][-1], 'Yes')  # Overdue column
        self.assertEqual(rows[1][-2], '7000.00')  # Balance due

    def test_receptionist_is_forbidden(self):
        self.authenticate(self.receptionist)
        response = self.client.get(reverse('export-outstanding-dues'))
        self.assertEqual(response.status_code, 403)


class OccupancyExportTests(ExportTestCase):
    def setUp(self):
        super().setUp()
        self.bed_b = self.create_bed(self.room, bed_number='101-B')
        self.bed_c = self.create_bed(self.room, bed_number='101-C')
        with tenant_context(self.tenant.id):
            self.bed_c.status = Bed.Status.MAINTENANCE
            self.bed_c.save()
        resident = self.create_resident(self.property, status=Resident.Status.RESERVED)
        self.check_in(resident, self.bed)  # occupies self.bed

    def test_counts_beds_by_status(self):
        rows = _csv_rows(self.client.get(reverse('export-occupancy')))

        self.assertEqual(rows[0], ['Property', 'Total Beds', 'Occupied', 'Available', 'Reserved', 'Maintenance', 'Occupancy %'])
        data_row = next(r for r in rows[1:] if r[0] == 'Sunrise PG')
        self.assertEqual(data_row[1], '3')  # total
        self.assertEqual(data_row[2], '1')  # occupied
        self.assertEqual(data_row[3], '1')  # available (bed_b)
        self.assertEqual(data_row[5], '1')  # maintenance (bed_c)

    def test_property_filter_narrows_to_one_property(self):
        other_property = self.create_property(self.tenant, name='Other PG')
        other_floor = self.create_floor(other_property)
        other_room = self.create_room(other_floor, room_number='301')
        self.create_bed(other_room, bed_number='301-A')

        rows = _csv_rows(
            self.client.get(reverse('export-occupancy'), {'property': str(self.property.id)})
        )
        names = [r[0] for r in rows[1:]]
        self.assertEqual(names, ['Sunrise PG'])

    def test_receptionist_is_forbidden(self):
        self.authenticate(self.receptionist)
        response = self.client.get(reverse('export-occupancy'))
        self.assertEqual(response.status_code, 403)
